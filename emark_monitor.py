import streamlit as st
import psutil
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import uiautomation as auto

# =====================
# Excel setup
# =====================
EXCEL_FILE = "emark_downtime.xlsx"

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Start Time", "End Time", "Duration (minutes)", "Reason"])
        wb.save(EXCEL_FILE)

def log_excel(start, end, reason):
    duration = (end - start).total_seconds() / 60
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([str(start), str(end), round(duration, 2), reason])
    wb.save(EXCEL_FILE)

init_excel()

# =====================
# Detection functions
# =====================
def get_emark_proc():
    for proc in psutil.process_iter(['name']):
        if proc.info['name'] and "emark" in proc.info['name'].lower():
            return proc
    return None

def is_frozen(proc):
    try:
        return proc.status() == psutil.STATUS_NOT_RESPONDING
    except:
        return True

def get_button_states():
    window = None
    for w in auto.GetRootControl().GetChildren():
        if "emark" in w.Name.lower():
            window = w
            break

    if window is None:
        return "DOWN", "UI not found"

    start_btn = window.ButtonControl(Name="Start")
    stop_btn  = window.ButtonControl(Name="Stop")

    if not start_btn.Exists(0,0) or not stop_btn.Exists(0,0):
        return "DOWN", "Buttons not found"

    if start_btn.IsEnabled and not stop_btn.IsEnabled:
        return "STOPPED", "Start enabled, Stop disabled"
    if stop_btn.IsEnabled and not start_btn.IsEnabled:
        return "RUNNING", "Stop enabled, Start disabled"

    return "DOWN", "Unexpected UI state"

# =====================
# Streamlit UI
# =====================
st.title("🟢 eMark Monitoring System")

# Session state to track downtime
if "is_down" not in st.session_state:
    st.session_state.is_down = False
if "down_start" not in st.session_state:
    st.session_state.down_start = None
if "reason" not in st.session_state:
    st.session_state.reason = ""

status_box = st.empty()
log_box = st.empty()

# =====================
# Detection logic
# =====================
proc = get_emark_proc()

if proc is None:
    state = "DOWN"
    reason = "Process closed"
elif is_frozen(proc):
    state = "DOWN"
    reason = "Application frozen"
else:
    state, reason = get_button_states()

status_box.write(f"**Status:** `{state}` — {reason}")

# Detect DOWN start
if state != "RUNNING" and not st.session_state.is_down:
    st.session_state.is_down = True
    st.session_state.down_start = datetime.now()
    st.session_state.reason = reason
    log_box.write(f"🔴 DOWN since {str(st.session_state.down_start)} — {reason}")

# Detect DOWN end
if state == "RUNNING" and st.session_state.is_down:
    st.session_state.is_down = False
    end_time = datetime.now()
    log_excel(st.session_state.down_start, end_time, st.session_state.reason)
    log_box.write(f"🟢 UP at {str(end_time)} — downtime logged")

# =====================
# Auto-refresh every 1 second
# =====================
st_autorefresh = st.experimental_data_editor if hasattr(st, "experimental_data_editor") else st.empty()
st_autorefresh = st_autorefresh  # dummy placeholder

# Streamlit built-in auto-refresh
st.experimental_set_query_params(_t=str(datetime.now()))
